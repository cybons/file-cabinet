import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def test_export(data):
    with pd.ExcelWriter("test.xlsx", engine="openpyxl") as writer:
        data.to_excel(writer, sheet_name="全比較", index=False)


def export_excel_data_formatting(path, results_df):
    """エクスポートデータ前処理"""

    # 未確定のデータのみを抽出
    undetermined_df = results_df[~results_df["確定"]]

    # ランク差が2以下の組織のみをフィルタリング
    filtered_by_rank_df = undetermined_df[undetermined_df["ランク差"] <= 2]

    # ランク差に基づいてソート
    filtered_by_rank_df = filtered_by_rank_df.sort_values(by=["ランク差", "前月の組織"])

    # Excelに出力

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        results_df.to_excel(writer, sheet_name="全比較", index=False)
        filtered_by_rank_df.to_excel(
            writer, sheet_name="未確定のみ（ランク差上位2）", index=False
        )

    # ワークブックを開いてテーブルを追加
    wb = load_workbook("D:\\Work\\python\\comparison_results.xlsx")
    ws = wb["全比較"]

    # テーブルの範囲を指定
    data_range = (
        f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"  # noqa: E501
    )
    table = Table(displayName="Table1", ref=data_range)

    # テーブルスタイルを適用 (Excel に組み込みのスタイル名)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # セル幅を調整
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # カラムの文字を取得
        for cell in col:
            try:  # セルが数値の場合、len関数がエラーを出すので文字列に変換
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(writer)
