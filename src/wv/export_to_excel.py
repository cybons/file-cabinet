"""
結果データをExcelファイルにエクスポートするモジュール。
"""

import constants as cst
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


def export_to_excel(df_results: pd.DataFrame, file_path: str) -> None:
    """
    結果データをExcelファイルにエクスポートする関数。

    Args:
        df_results (pd.DataFrame): 結果データフレーム。
        file_path (str): エクスポート先のファイルパス。
    """
    # ワークブックの作成
    wb = Workbook()

    # 全体シートの作成
    ws_all = wb.active
    ws_all.title = "全体"
    for r in dataframe_to_rows(df_results, index=False, header=True):
        ws_all.append(r)
    create_table_and_style(ws_all, "全体")

    # 未確定シートの作成
    df_unconfirmed = df_results[~df_results[cst.CONFIRMED]]
    ws_unconfirmed = wb.create_sheet(title="未確定")

    for r in dataframe_to_rows(df_unconfirmed, index=False, header=True):
        ws_unconfirmed.append(r)
    create_table_and_style(ws_unconfirmed, "未確定")

    # ファイルに保存
    wb.save(file_path)


def create_table_and_style(ws, title: str) -> None:
    """
    Excelシートにテーブルとスタイルを適用するヘルパー関数。

    Args:
        ws (Worksheet): 対象のワークシート。
        title (str): テーブルのタイトル。
    """
    # テーブル範囲の設定
    table_range = f"A1:{chr(64 + ws.max_column)}{ws.max_row}"
    table = Table(displayName=title, ref=table_range)

    # テーブルスタイルの設定
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # 見出しを縦書きに設定
    for cell in ws["1:1"]:
        cell.alignment = Alignment(textRotation=255)

    # ウィンドウ枠の固定
    ws.freeze_panes = ws["C2"]
