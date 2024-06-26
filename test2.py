import os
from collections import defaultdict
from itertools import product

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

import constants


class OrganizationMatcher:
    def __init__(self, config):
        self.config = config
        self.total_weight = (
            config["weights"][constants.RANK_WEIGHT]
            + config["weights"][constants.SIMILARITY_WEIGHT]
            + config["weights"][constants.MEMBER_WEIGHT]
        )

    def get_thresholds(self, size):
        thresholds = sorted(
            self.config[constants.THRESHOLDS], key=lambda x: x[constants.SIZE]
        )
        for threshold in thresholds:
            if size <= threshold[constants.SIZE]:
                return threshold
        return thresholds[-1]  # 最大サイズを超えた場合は最後の閾値を返す

    def calculate_rank(self, org):
        return org.count("/") + 1

    def calculate_rank_score(self, rank_diff):
        base_rank_weight = self.config["weights"][constants.RANK_WEIGHT]
        if rank_diff == 0:
            return base_rank_weight
        elif rank_diff > 0:
            # ランクダウンの場合、スコアはランク差に応じて減少（0.5ずつ減少）
            return (
                base_rank_weight - rank_diff * 0.5
                if base_rank_weight - rank_diff * 0.5 > 0
                else 0
            )
        else:
            # ランクアップの場合、スコアはランク差に応じてさらに減少（1ずつ減少）
            return (
                base_rank_weight - abs(rank_diff)
                if base_rank_weight - abs(rank_diff) > 0
                else 0
            )

    def calculate_similarity_score_with_ratio(self, row_dict):
        size_a = row_dict[constants.PREV_SIZE]
        size_b = row_dict[constants.CURR_SIZE]
        thresholds_a = self.get_thresholds(size_a)
        thresholds_b = self.get_thresholds(size_b)

        similarity_threshold = min(
            thresholds_a[constants.SIMILARITY_THRESHOLD],
            thresholds_b[constants.SIMILARITY_THRESHOLD],
        )
        member_ratio_threshold = min(
            thresholds_a[constants.MEMBER_RATIO_THRESHOLD],
            thresholds_b[constants.MEMBER_RATIO_THRESHOLD],
        )

        # ランク差のスコアを設定
        rank_score = self.calculate_rank_score(row_dict[constants.RANK_DIFF])

        similarity_score = (
            row_dict[constants.SIMILARITY_INDEX] >= similarity_threshold
        ) * self.config["weights"][constants.SIMILARITY_WEIGHT]
        member_score = (
            row_dict[constants.COMMON_RATIO] >= member_ratio_threshold
        ) * self.config["weights"][constants.MEMBER_WEIGHT]
        total_score = rank_score + similarity_score + member_score
        same_org = total_score >= self.total_weight

        # 3人未満で同じ組織名の場合、同一組織と見なす
        if (
            size_a < 3
            and size_b
            and row_dict[constants.PREV_ORG] == row_dict[constants.CURR_ORG]
        ):
            same_org = True

        return {
            "ランクスコア": rank_score,
            "類似度スコア": similarity_score,
            "メンバースコア": member_score,
            "総合スコア": total_score,
            "適用ルール": thresholds_a["comment"],
            constants.SAME_ORG: same_org,
        }


def load_config():
    path = os.path.dirname(__file__)
    filepath = os.path.join(path, "organization_matching_config.yaml")
    with open(filepath, "r") as file:
        config = yaml.safe_load(file)
    return config


# データフレームから組織とユーザーのマッピングをdictとして作成する関数
def build_org_user_set(df, additional_columns=None):
    if additional_columns is None:
        additional_columns = []

    org_users = defaultdict(lambda: {"users": set(), "info": defaultdict(list)})
    for record in df.to_dict("records"):
        org_name = record[constants.ORG]

        # 「組織」という最上位組織を除外
        if org_name == "組織":
            continue

        org_hierarchy = org_name.split("/")

        # 上位の組織にユーザーを追加
        for i in range(len(org_hierarchy)):
            current_org = "/".join(org_hierarchy[: i + 1])
            if current_org == "組織":
                continue
            org_users[current_org]["users"].add(record[constants.USER])
            for col in additional_columns:
                org_users[current_org]["info"][col].append(record[col])

    return org_users


# 社員と派遣社員の比率を計算する関数
def calculate_employee_ratio(info):
    total = len(info[constants.TYPE])
    full_time = info[constants.TYPE].count(constants.FULL_TIME)
    part_time = info[constants.TYPE].count(constants.PART_TIME)
    contract_employee = info[constants.TYPE].count(constants.CONTRACT_EMPLOYEE)
    return {
        "full_time_ratio": full_time / total if total > 0 else 0,
        "part_time_ratio": part_time / total if total > 0 else 0,
        "contract_employee_ratio": contract_employee / total if total > 0 else 0,
    }


# 組織の構成比率の差を計算する関数
def calculate_ratio_difference(ratio_a, ratio_b):
    full_time_diff = abs(ratio_a["full_time_ratio"] - ratio_b["full_time_ratio"])
    part_time_diff = abs(ratio_a["part_time_ratio"] - ratio_b["part_time_ratio"])
    contract_employee_diff = abs(
        ratio_a["contract_employee_ratio"] - ratio_b["contract_employee_ratio"]
    )

    # 各比率の差の平均を計算し、100から引いて100%表記の似ている度合いを示す
    avg_diff = (full_time_diff + part_time_diff + contract_employee_diff) / 3
    similarity_percentage = 1 - avg_diff
    return similarity_percentage


# 組織ランクを計算する関数
def calculate_rank(org):
    return org.count("/") + 1


# ランク差に基づくスコアを計算する関数
def calculate_rank_score(rank_diff, base_rank_weight):
    if rank_diff == 0:
        return base_rank_weight
    elif rank_diff > 0:
        # ランクダウンの場合、スコアはランク差に応じて減少（0.5ずつ減少）
        return (
            base_rank_weight - rank_diff * 0.5
            if base_rank_weight - rank_diff * 0.5 > 0
            else 0
        )
    else:
        # ランクアップの場合、スコアはランク差に応じてさらに減少（1ずつ減少）
        return (
            base_rank_weight - abs(rank_diff)
            if base_rank_weight - abs(rank_diff) > 0
            else 0
        )


# 組織の一致判定と確定処理を行う関数
def update_organization_matches(df_results):
    df_results[constants.CONFIRMED] = False

    # 同一組織と見なされている場合に一括でCONFIRMEDをTrueに更新
    df_results.loc[df_results[constants.SAME_ORG], constants.CONFIRMED] = True

    # 確定がTrueになった前月組織と当月組織の同じ名前のものをさらに更新
    confirmed_prev = df_results[df_results[constants.CONFIRMED]][
        constants.PREV_ORG
    ].unique()
    confirmed_curr = df_results[df_results[constants.CONFIRMED]][
        constants.CURR_ORG
    ].unique()

    df_results.loc[
        df_results[constants.PREV_ORG].isin(confirmed_prev), constants.CONFIRMED
    ] = True
    df_results.loc[
        df_results[constants.CURR_ORG].isin(confirmed_curr), constants.CONFIRMED
    ] = True

    return df_results


data_A = [
    {"org": "営業部", "user": "u1", "type": "full_time"},
    {"org": "営業部", "user": "u2", "type": "full_time"},
    {"org": "営業部", "user": "u3", "type": "full_time"},
    {"org": "営業部/営業課", "user": "u4", "type": "part_time"},
    {"org": "営業部/営業課", "user": "u5", "type": "part_time"},
    {"org": "営業部/営業課", "user": "u6", "type": "part_time"},
    {"org": "開発部", "user": "u7", "type": "full_time"},
    {"org": "開発部", "user": "u8", "type": "full_time"},
    {"org": "開発部", "user": "u9", "type": "full_time"},
    {"org": "開発部/開発課", "user": "u10", "type": "part_time"},
    {"org": "開発部/開発課", "user": "u11", "type": "part_time"},
    {"org": "開発部/開発課", "user": "u12", "type": "part_time"},
    {"org": "サポート部", "user": "u13", "type": "full_time"},
    {"org": "サポート部", "user": "u14", "type": "full_time"},
    {"org": "サポート部", "user": "u15", "type": "full_time"},
    {"org": "サポート部/サポート課", "user": "u16", "type": "part_time"},
    {"org": "サポート部/サポート課", "user": "u17", "type": "part_time"},
    {"org": "サポート部/サポート課", "user": "u18", "type": "part_time"},
    {"org": "企画部", "user": "u19", "type": "full_time"},
    {"org": "企画部", "user": "u20", "type": "full_time"},
    {"org": "企画部", "user": "u21", "type": "full_time"},
    {"org": "企画部/企画課", "user": "u22", "type": "part_time"},
    {"org": "企画部/企画課", "user": "u23", "type": "part_time"},
    {"org": "企画部/企画課", "user": "u24", "type": "part_time"},
    {"org": "人事部", "user": "u25", "type": "full_time"},
    {"org": "人事部", "user": "u26", "type": "full_time"},
    {"org": "人事部", "user": "u27", "type": "full_time"},
    {"org": "人事部/人事課", "user": "u28", "type": "part_time"},
    {"org": "人事部/人事課", "user": "u29", "type": "part_time"},
    {"org": "人事部/人事課", "user": "u30", "type": "part_time"},
    {"org": "法務部", "user": "u31", "type": "full_time"},
    {"org": "法務部", "user": "u32", "type": "full_time"},
    {"org": "法務部", "user": "u33", "type": "full_time"},
    {"org": "法務部/法務課", "user": "u34", "type": "part_time"},
    {"org": "法務部/法務課", "user": "u35", "type": "part_time"},
    {"org": "法務部/法務課", "user": "u36", "type": "part_time"},
    {"org": "経理部", "user": "u37", "type": "full_time"},
    {"org": "経理部", "user": "u38", "type": "full_time"},
    {"org": "経理部", "user": "u39", "type": "full_time"},
    {"org": "経理部/経理課", "user": "u40", "type": "part_time"},
    {"org": "経理部/経理課", "user": "u41", "type": "part_time"},
    {"org": "経理部/経理課", "user": "u42", "type": "part_time"},
]


df_A = pd.DataFrame(data_A)

data_B = [
    {"org": "営業部", "user": "u1", "type": "full_time"},
    {"org": "営業部", "user": "u2", "type": "full_time"},
    {"org": "営業部", "user": "u3", "type": "full_time"},
    {"org": "営業部/営業課", "user": "u4", "type": "part_time"},
    {"org": "営業部/営業課", "user": "u5", "type": "part_time"},
    {"org": "営業部/営業課", "user": "u6", "type": "part_time"},
    {"org": "開発部", "user": "u9", "type": "full_time"},
    {"org": "開発部", "user": "u10", "type": "full_time"},
    {"org": "開発部", "user": "u11", "type": "full_time"},
    {"org": "開発部/開発課", "user": "u12", "type": "part_time"},
    {"org": "開発部/開発課", "user": "u13", "type": "part_time"},
    {"org": "開発部/開発課", "user": "u14", "type": "part_time"},
    {"org": "サポート部", "user": "u15", "type": "full_time"},
    {"org": "サポート部", "user": "u16", "type": "full_time"},
    {"org": "サポート部", "user": "u17", "type": "full_time"},
    {"org": "サポート部/サポート課", "user": "u18", "type": "part_time"},
    {"org": "サポート部/サポート課", "user": "u19", "type": "part_time"},
    {"org": "サポート部/サポート課", "user": "u20", "type": "part_time"},
    {"org": "企画部", "user": "u21", "type": "full_time"},
    {"org": "企画部", "user": "u22", "type": "full_time"},
    {"org": "企画部", "user": "u23", "type": "full_time"},
    {"org": "企画部/企画課", "user": "u24", "type": "part_time"},
    {"org": "企画部/企画課", "user": "u25", "type": "part_time"},
    {"org": "企画部/企画課", "user": "u26", "type": "part_time"},
    {"org": "人事部", "user": "u27", "type": "full_time"},
    {"org": "人事部", "user": "u28", "type": "full_time"},
    {"org": "人事部", "user": "u29", "type": "full_time"},
    {"org": "人事部/人事課", "user": "u30", "type": "part_time"},
    {"org": "人事部/人事課", "user": "u31", "type": "part_time"},
    {"org": "人事部/人事課", "user": "u32", "type": "part_time"},
    {"org": "法務部", "user": "u33", "type": "full_time"},
    {"org": "法務部", "user": "u34", "type": "full_time"},
    {"org": "法務部", "user": "u35", "type": "full_time"},
    {"org": "法務部/法務課", "user": "u36", "type": "part_time"},
    {"org": "法務部/法務課", "user": "u37", "type": "part_time"},
    {"org": "法務部/法務課", "user": "u38", "type": "part_time"},
    {"org": "経理部", "user": "u39", "type": "full_time"},
    {"org": "経理部", "user": "u40", "type": "full_time"},
    {"org": "経理部", "user": "u41", "type": "full_time"},
    {"org": "経理部/経理課", "user": "u42", "type": "part_time"},
    {"org": "経理部/経理課", "user": "u43", "type": "part_time"},
    {"org": "経理部/経理課", "user": "u44", "type": "part_time"},
]
df_B = pd.DataFrame(data_B)
# データの読み込みおよび準備（例: df_A, df_B、rank_diff）


def process_organization_pair(org_a, org_b, matcher):
    org_a_name, data_A = org_a
    org_b_name, data_B = org_b
    users_A = data_A["users"]
    users_B = data_B["users"]
    size_a = len(users_A)
    size_b = len(users_B)
    intersection = users_A.intersection(users_B)
    if not intersection:
        return None

    union = users_A.union(users_B)
    intersection_size = len(intersection)
    union_size = len(union)
    ratio_a = calculate_employee_ratio(data_A["info"])
    ratio_b = calculate_employee_ratio(data_B["info"])
    ratio_diff = calculate_ratio_difference(ratio_a, ratio_b)
    rank_diff = calculate_rank(org_a_name) - calculate_rank(org_b_name)
    result = {
        constants.PREV_ORG: org_a_name,
        constants.CURR_ORG: org_b_name,
        constants.PREV_SIZE: size_a,
        constants.CURR_SIZE: size_b,
        constants.COMMON_MEMBERS: intersection_size,
        constants.COMMON_RATIO: intersection_size / size_b if size_b > 0 else 0,
        constants.SIMILARITY_INDEX: intersection_size / union_size
        if union_size > 0
        else 0,
        constants.RATIO_DIFF: ratio_diff,
        constants.RANK_DIFF: rank_diff,
    }

    similarity_scores = matcher.calculate_similarity_score_with_ratio(result)
    result.update(similarity_scores)
    return result


def calculate_and_update_organization_scores(df_A, df_B):
    # データフレームから組織とユーザーのマッピングをdictとして作成
    org_users_A = build_org_user_set(df_A, additional_columns=[constants.TYPE])
    org_users_B = build_org_user_set(df_B, additional_columns=[constants.TYPE])

    config = load_config()
    matcher = OrganizationMatcher(config)

    # 結果リストを作成
    results = []
    for org_a, org_b in product(org_users_A.items(), org_users_B.items()):
        result = process_organization_pair(org_a, org_b, matcher)
        if result:
            results.append(result)

    # スコアを計算して各列に分解して代入
    df_results = pd.DataFrame(results)

    # 組織の一致判定と確定処理
    df_results = update_organization_matches(df_results)

    return df_results


def export_to_excel(df_results, filename):
    wb = Workbook()

    # 全体シートの作成
    ws_all = wb.active
    ws_all.title = "全体"

    for r_idx, row in enumerate(
        dataframe_to_rows(df_results, index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            cell = ws_all.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                # ヘッダー行を縦書きに設定
                cell.alignment = Alignment(textRotation=255)

    # 確定がFalseのものを未確定シートに書き込む
    df_unconfirmed = df_results[~df_results[constants.CONFIRMED]]
    ws_unconfirmed = wb.create_sheet(title="未確定")

    for r_idx, row in enumerate(
        dataframe_to_rows(df_unconfirmed, index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            cell = ws_unconfirmed.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                # ヘッダー行を縦書きに設定
                cell.alignment = Alignment(textRotation=255)

    # テーブルを作成してシートに適用
    def create_table(ws, df, table_name):
        table = Table(displayName=table_name, ref=ws.dimensions)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    create_table(ws_all, df_results, "AllDataTable")
    create_table(ws_unconfirmed, df_unconfirmed, "UnconfirmedDataTable")

    # ウィンドウ枠の固定
    ws_all.freeze_panes = "C2"  # 1行目を固定
    ws_unconfirmed.freeze_panes = "C2"  # 1行目を固定

    # Excelファイルに保存
    wb.save(filename)


df_results = calculate_and_update_organization_scores(df_A, df_B)


def find_disjoint_organizations(results, df_A, df_B):
    # 結果から前月と当月の組織を抽出
    prev_orgs_in_results = set(results[constants.PREV_ORG])
    curr_orgs_in_results = set(results[constants.CURR_ORG])

    # 前月と当月の全組織を取得
    all_prev_orgs = set(df_A[constants.ORG])
    all_curr_orgs = set(df_B[constants.ORG])

    # 共通ユーザーがいない組織を特定
    disjoint_prev_orgs = all_prev_orgs - prev_orgs_in_results
    disjoint_curr_orgs = all_curr_orgs - curr_orgs_in_results

    # 結果をデータフレームに変換
    disjoint_df = pd.DataFrame(
        {
            "組織": list(disjoint_prev_orgs) + list(disjoint_curr_orgs),
            "状態": ["抹消"] * len(disjoint_prev_orgs)
            + ["新設"] * len(disjoint_curr_orgs),
        }
    )

    return disjoint_df


# 結果を表示
print(df_results)

export_to_excel(df_results, "pandas_to_excel.xlsx")


def check_results_integrity(file_path, sheet_name="未確定"):
    """
    ファイルの未確定シートを読み込み、データの整合性をチェックします。
    """
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # 確認列が含まれているかチェック
    if "確認" not in df.columns:
        raise ValueError("確認列が見つかりません。ファイルが正しいか確認してください。")

    # 確認列が "⚪︎" または "-" のみ含んでいるかチェック
    if not df["確認"].isin(["⚪︎", "-"]).all():
        raise ValueError("確認列に不正な値が含まれています。")

    # 確認列が "⚪︎" の前月組織が一意かチェック
    duplicate_prev_orgs = df[df["確認"] == "⚪︎"][constants.PREV_ORG].duplicated().any()
    if duplicate_prev_orgs:
        raise ValueError("確認列において前月組織が重複しています。")

    # 確認列が "⚪︎" の当月組織が一意かチェック
    duplicate_curr_orgs = df[df["確認"] == "⚪︎"][constants.CURR_ORG].duplicated().any()
    if duplicate_curr_orgs:
        raise ValueError("確認列において当月組織が重複しています。")

    return df


def update_results_with_confirmation(df_results, checked_df):
    """
    チェックされたデータフレームに基づいてresultsを更新します。
    """
    # 確認列が "⚪︎" の行を取得
    confirmed_rows = checked_df[checked_df["確認"] == "⚪︎"]

    # 既存の確定データを一旦未確定にして同一組織もFalseにする（確認が "⚪︎" になった組織のみ）
    for _, row in confirmed_rows.iterrows():
        prev_org = row[constants.PREV_ORG]
        curr_org = row[constants.CURR_ORG]

        # 該当する組織ペアを未確定にして同一組織をFalseに設定
        df_results.loc[
            (df_results[constants.PREV_ORG] == prev_org)
            | (df_results[constants.CURR_ORG] == curr_org),
            [constants.SAME_ORG, constants.CONFIRMED],
        ] = False

    # confirmed_rowsの同一組織と判定された組織ペアをresultsに反映
    for _, row in confirmed_rows.iterrows():
        prev_org = row[constants.PREV_ORG]
        curr_org = row[constants.CURR_ORG]

        # 該当する組織ペアを確定と同一組織に設定
        df_results.loc[
            (df_results[constants.PREV_ORG] == prev_org)
            & (df_results[constants.CURR_ORG] == curr_org),
            [constants.SAME_ORG, constants.CONFIRMED],
        ] = True

    # 同名組織の確定を再度行う
    for i, row in df_results.iterrows():
        if row[constants.CONFIRMED]:
            prev_org = row[constants.PREV_ORG]
            curr_org = row[constants.CURR_ORG]
            df_results.loc[
                (df_results[constants.PREV_ORG] == prev_org)
                | (df_results[constants.CURR_ORG] == curr_org),
                constants.CONFIRMED,
            ] = True

    return df_results


class SimulationOrg:
    def __init__(self, org_name, group_id=None, rank=None, parent=None):
        self.org_name = org_name
        self.group_id = group_id
        self.rank = rank
        self.parent = parent
        self.children = []

    def add_child(self, child_org):
        self.children.append(child_org)
        child_org.parent = self

    def update_org_name(self, new_name):
        self.org_name = new_name
        for child in self.children:
            child.update_tree_name()

    def update_tree_name(self):
        if self.parent:
            self.org_name = f"{self.parent.get_tree_name().split('/')[-1]}/{self.org_name.split('/')[-1]}"
            for child in self.children:
                child.update_tree_name()

    def get_tree_name(self):
        if self.parent:
            return f"{self.parent.get_tree_name()}/{self.org_name.split('/')[-1]}"
        return self.org_name

    def __repr__(self):
        return f"SimulationOrg(org_name={self.org_name}, group_id={self.group_id}, rank={self.rank})"


def simulate_updates_with_hierarchy(prev_orgs, curr_orgs, confirmed_df):
    # ランク別辞書を作成
    rank_dict = defaultdict(list)

    # 前月組織をランク別に分類
    for _, row in prev_orgs.iterrows():
        org = SimulationOrg(
            org_name=row["org"], group_id=row["group_id"], rank=row["rank"]
        )
        rank_dict[row["rank"]].append(org)

    # 当月組織をランク別に分類
    for _, row in curr_orgs.iterrows():
        org = SimulationOrg(org_name=row["org"], rank=row["rank"])
        rank_dict[row["rank"]].append(org)

    # 親子関係の構築
    for rank in sorted(rank_dict.keys()):
        for org in rank_dict[rank]:
            parent_name = "/".join(org.org_name.split("/")[:-1])
            if parent_name:
                for parent_org in rank_dict[rank - 1]:
                    if parent_org.org_name.split("/")[-1] == parent_name:
                        parent_org.add_child(org)
                        break

    # ツリーネームリストの初期化
    tree_name_list = {
        org.get_tree_name() for rank in rank_dict for org in rank_dict[rank]
    }

    # 確定データをもとに同一組織の更新シミュレーションを行う
    updates = []
    for _, row in confirmed_df.iterrows():
        if row[constants.SAME_ORG]:
            prev_org_name = row[constants.PREV_ORG]
            curr_org_name = row[constants.CURR_ORG]
            for org in rank_dict[row["rank_prev"]]:
                if org.org_name == prev_org_name:
                    if curr_org_name not in tree_name_list:
                        org.update_org_name(curr_org_name)
                        tree_name_list = {
                            org.get_tree_name()
                            for rank in rank_dict
                            for org in rank_dict[rank]
                        }
                    else:
                        # 一時的な名前に変更してから更新する
                        temp_name = f"{curr_org_name}_temp"
                        org.update_org_name(temp_name)
                        tree_name_list = {
                            org.get_tree_name()
                            for rank in rank_dict
                            for org in rank_dict[rank]
                        }
                        org.update_org_name(curr_org_name)
                        tree_name_list = {
                            org.get_tree_name()
                            for rank in rank_dict
                            for org in rank_dict[rank]
                        }

    # 更新データの作成
    for rank in sorted(rank_dict.keys()):
        for org in rank_dict[rank]:
            if org.group_id:
                tree_name = org.get_tree_name()
                updates.append({"group_id": org.group_id, "org_name": tree_name})

    return pd.DataFrame(updates)
